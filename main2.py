import openpyxl
from openpyxl import load_workbook
from datetime import date

# Load the existing workbook
workbook = load_workbook("attendance.xlsx")

# Select the active sheet
sheet = workbook.active

# Get today's date
today = date.today()

# Input student name and attendance status
student_name = input("Enter student name: ")
attendance_status = input("Enter attendance status (Present/Absent): ")

# Find the first empty row in the sheet
row_number = sheet.max_row + 1
