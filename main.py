import openpyxl
from openpyxl import load_workbook
from datetime import date
import tkinter as tk
from tkinter import simpledialog, messagebox

# Function to input student name and attendance status using a dialog
def input_attendance():
    student_name = simpledialog.askstring("Attendance Tracker", "Enter student name (or 'done' to finish):")
    if student_name is None or student_name.lower() == 'done':
        return None, None
    attendance_status = simpledialog.askstring("Attendance Tracker", "Enter attendance status (Present/Absent):")
    return student_name, attendance_status
# Function to load and display previous attendance records
def load_previous_attendance():
    try:
        workbook = load_workbook("attendance.xlsx")
        sheet = workbook.active
        for row in sheet.iter_rows(min_row=2, values_only=True):
            date, student_name, attendance_status = row
            attendance_listbox.insert(tk.END, f"Date: {date}, Student: {student_name}, Status: {attendance_status}")
        workbook.close()
    except FileNotFoundError:
        messagebox.showinfo("Info", "No previous attendance records found.")

# Generate a unique filename based on the current date and time
today = date.today()
file_suffix = today.strftime("%Y%m%d")
file_name = f"attendance_{file_suffix}.xlsx"

# Create a new Excel workbook if it doesn't exist
try:
    workbook = load_workbook("attendance.xlsx")
except FileNotFoundError:
    workbook = openpyxl.Workbook()
    workbook.active.append(["Date", "Student Name", "Attendance Status"])
    workbook.save("attendance.xlsx")

# Select the active sheet
sheet = workbook.active

# Create a tkinter window (GUI)
root = tk.Tk()
root.title("Attendance Tracker")

# Create a listbox to display attendance records
attendance_listbox = tk.Listbox(root)
attendance_listbox.pack(fill=tk.BOTH, expand=True)

# Load previous attendance records
load_previous_attendance()

# Input attendance using dialogs until the user is finished
while True:
    student_name, attendance_status = input_attendance()
    if student_name is None:
        break

    # Find the first empty row in the sheet
    row_number = sheet.max_row + 1

    # Populate the sheet with data
    sheet[f"A{row_number}"] = today.strftime("%Y-%m-%d")
    sheet[f"B{row_number}"] = student_name
    sheet[f"C{row_number}"] = attendance_status

    # Update the attendance listbox
    attendance_listbox.insert(tk.END, f"Date: {today.strftime('%Y-%m-%d')}, Student: {student_name}, Status: {attendance_status}")

# Save the workbook
workbook.save("attendance.xlsx")

messagebox.showinfo("Info", f"Attendance recorded successfully.")

# Run the tkinter main loop
root.mainloop()
